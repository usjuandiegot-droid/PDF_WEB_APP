#INCAPACIDADES V2
import fitz  # PyMuPDF
import pandas as pd
import re
import os
import json
import traceback

from datetime import datetime
from zoneinfo import ZoneInfo
from io import BytesIO

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image


# ==========================================================
# FLASK
# ==========================================================

app = Flask(__name__)

CORS(
    app,
    expose_headers=["Content-Disposition"]
)

app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024


# ==========================================================
# LOGGER
# ==========================================================

class JsonLogger:

    def log(self, level, event, message, extra=None):

        log = {
            "timestamp": datetime.now(
                ZoneInfo("America/Bogota")
            ).isoformat(),
            "level": level,
            "event": event,
            "message": message
        }

        if extra:
            log["extra"] = extra

        print(json.dumps(log))

    def info(self, event, message, extra=None):
        self.log("info", event, message, extra)

    def error(self, event, message, extra=None):
        self.log("error", event, message, extra)


logger = JsonLogger()


# ==========================================================
# ARCHIVOS BASE
# ==========================================================

BASE_DIR = os.path.dirname(
    os.path.abspath(__file__)
)

PLANTILLA_FILE = os.path.join(
    BASE_DIR,
    "PLANTILLA.xlsx"
)

CIE10_FILE = os.path.join(
    BASE_DIR,
    "CIE10.xlsx"
)

FIRMA_FILE = os.path.join(
    BASE_DIR,
    "firma.png"
)


# ==========================================================
# CARGAR CIE10
# ==========================================================

cie10_df = pd.read_excel(CIE10_FILE)

cie10_df["Codigo"] = (
    cie10_df["Codigo"]
    .astype(str)
    .str.strip()
)


# ==========================================================
# FUNCIONES
# ==========================================================

def calculate_total_days(start_date_str, end_date_str):

    if start_date_str is None or end_date_str is None:
        return None

    try:

        start_date = datetime.strptime(
            start_date_str,
            "%d/%m/%Y"
        )

        end_date = datetime.strptime(
            end_date_str,
            "%d/%m/%Y"
        )

        return (
            end_date - start_date
        ).days + 1

    except ValueError:

        return None


# ==========================================================
# EXTRACCIÓN DE INFORMACIÓN DEL PDF
# (Basada exactamente en el script de Colab)
# ==========================================================

def extract_data(text):

    # Inicializar variables
    patient_name = None
    identification = None
    diagnosis = None
    date_of_attention = None
    order = None
    prorogation = None
    start_date = None
    end_date = None
    eps = None
    tipo_incapacidad = None
    grupo_servicio = None

    # ======================================================
    # PACIENTE
    # ======================================================

    patient_name_match = re.search(
        r'PRESCRIPCIÓN DE INCAPACIDAD/LICENCIA DE MATERNIDAD\s*([^\n]*)\s*Identificación:',
        text
    )

    if patient_name_match:
        patient_name = patient_name_match.group(1).strip()

    # ======================================================
    # DOCUMENTO
    # ======================================================

    identification_match = re.search(
        r'Identificación:\s*CC\s*(\d+)',
        text
    )

    if identification_match:
        identification = identification_match.group(1).strip()

    # ======================================================
    # DIAGNÓSTICO
    # ======================================================

    diagnosis_code_match = re.search(
        r'\b([A-Z]\d{2,3}[A-Z]?)\b',
        text
    )

    if diagnosis_code_match:
        diagnosis = diagnosis_code_match.group(1).strip()
    else:
        diagnosis = None

    # ======================================================
    # EPS
    # ======================================================

    eps_lines = re.findall(
        r'EPS:\s*(.*)',
        text
    )

    if eps_lines:
        eps = eps_lines[-1].strip()

    # ======================================================
    # FECHA INICIO
    # ======================================================

    start_date_match = re.search(
        r'Fecha Inicio:\s*(\d{2}/\d{2}/\d{4})',
        text
    )

    if start_date_match:
        start_date = start_date_match.group(1).strip()

    # ======================================================
    # FECHA FIN
    # ======================================================

    end_date_match = re.search(
        r'Fecha Fin:\s*(\d{2}/\d{2}/\d{4})',
        text
    )

    if end_date_match:
        end_date = end_date_match.group(1).strip()

    # ======================================================
    # FECHA ATENCIÓN
    # ======================================================

    date_of_attention_match = re.search(
        r'Orden:\s*\d+\n(\d{4}/\d{2}/\d{2})',
        text
    )

    if date_of_attention_match:
        date_of_attention = date_of_attention_match.group(1).strip()

    # ======================================================
    # ORDEN
    # ======================================================

    order_match = re.search(
        r'Orden:\s*(\d+)',
        text
    )

    if order_match:
        order = order_match.group(1).strip()

    # ======================================================
    # PRÓRROGA
    # ======================================================

    prorogation_match = re.search(
        r'Prórroga:\s*(NO|SI)',
        text
    )

    if prorogation_match:
        prorogation = prorogation_match.group(1).strip()

    # ======================================================
    # TIPO INCAPACIDAD
    # ======================================================

    tipo_incapacidad_match = re.search(
        r'Tipo Incapacidad:\s*(.*)',
        text
    )

    if tipo_incapacidad_match:
        tipo_incapacidad = tipo_incapacidad_match.group(1).strip()

    # ======================================================
    # GRUPO SERVICIO
    # ======================================================

    grupo_servicio_match = re.search(
        r'Grupo Servicio:\s*(.*)',
        text
    )

    if grupo_servicio_match:
        grupo_servicio = grupo_servicio_match.group(1).strip()

    # ======================================================
    # TOTAL DÍAS
    # ======================================================

    total_days = calculate_total_days(
        start_date,
        end_date
    )

    # ======================================================
    # RETORNAR RESULTADO
    # ======================================================

    return {

        "Documento": identification,

        "Paciente": patient_name,

        "CONCATENAR":
            f"{identification} {patient_name}"
            if identification and patient_name
            else None,

        "Fecha de Inicio": start_date,

        "Fecha de Fin": end_date,

        "Total de Días": total_days,

        "Dias": None,

        "EPS": eps,

        "Observacion": None,

        "DX": diagnosis,

        "FECHA": date_of_attention,

        "ORDEN": order,

        "PRORROGA": prorogation,

        "Tipo Incapacidad": tipo_incapacidad,

        "Grupo Servicio": grupo_servicio
    }

# ==========================================================
# ENDPOINT
# ==========================================================

@app.route("/procesar", methods=["POST"])
def procesar():

    try:

        # ==========================================
        # RECIBIR ARCHIVOS
        # ==========================================

        files = request.files.getlist("files")

        if len(files) == 0:
            return jsonify({
                "error": "No se enviaron archivos."
            }), 400

        if len(files) > 29:
            return jsonify({
                "error": "Máximo 29 archivos por proceso."
            }), 400

        logger.info(
            "inicio",
            "Procesamiento iniciado",
            {
                "archivos": len(files)
            }
        )

        all_data = []

        # ==========================================
        # LEER CADA PDF
        # ==========================================

        for file in files:

            try:

                pdf_bytes = file.read()

                pdf_document = fitz.open(
                    stream=pdf_bytes,
                    filetype="pdf"
                )

                if len(pdf_document) == 0:
                    continue

                first_page = pdf_document[0]

                text = first_page.get_text("text")

                data = extract_data(text)

                all_data.append(data)

                pdf_document.close()

            except Exception as e:

                logger.error(
                    "pdf",
                    f"Error leyendo {file.filename}",
                    {
                        "error": str(e)
                    }
                )

        # ==========================================
        # VALIDAR RESULTADOS
        # ==========================================

        if len(all_data) == 0:

            return jsonify({
                "error": "No fue posible extraer información de ningún PDF."
            }), 400

        # ==========================================
        # DATAFRAME
        # ==========================================

        df = pd.DataFrame(all_data)

        # ==========================================
        # MERGE CIE10
        # (idéntico al script de Colab)
        # ==========================================

        df = pd.merge(
            df,
            cie10_df,
            left_on="DX",
            right_on="Codigo",
            how="left"
        )

        df["DX"] = (
            df["DX"].fillna("")
            + " "
            + df["Nombre"].fillna("")
        )

        df.drop(
            columns=[
                "Codigo",
                "Nombre"
            ],
            inplace=True
        )

        # ==========================================
        # FECHA ACTUAL
        # ==========================================

        fecha_actual = datetime.now(
            ZoneInfo("America/Bogota")
        )

        df.insert(
            0,
            "Fecha envio dra",
            fecha_actual.strftime("%d/%m/%Y")
        )

        # ==========================================
        # ORDENAR COLUMNAS
        # ==========================================

        columnas = [

            "Fecha envio dra",

            "Documento",

            "Paciente",

            "CONCATENAR",

            "Fecha de Inicio",

            "Fecha de Fin",

            "Total de Días",

            "Dias",

            "EPS",

            "Observacion",

            "DX",

            "FECHA",

            "ORDEN",

            "PRORROGA",

            "Tipo Incapacidad",

            "Grupo Servicio"

        ]

        df = df.reindex(columns=columnas)


        # ==========================================
        # CARGAR PLANTILLA
        # ==========================================

        wb = load_workbook(PLANTILLA_FILE)

        ws = wb["INFO"]

        # ==========================================
        # LIMPIAR DATOS ANTERIORES
        # ==========================================

        for row in ws.iter_rows(
            min_row=2,
            max_row=300,
            min_col=1,
            max_col=16
        ):
            for cell in row:
                cell.value = None

        # ==========================================
        # ESCRIBIR DATAFRAME
        # ==========================================

        for r_idx, row in enumerate(
            dataframe_to_rows(
                df,
                index=False,
                header=False
            ),
            start=2
        ):

            for c_idx, value in enumerate(
                row,
                start=1
            ):

                ws.cell(
                    row=r_idx,
                    column=c_idx
                ).value = value

        # ==========================================
        # INSERTAR FIRMAS
        # ==========================================

        if os.path.exists(FIRMA_FILE):

            fila_base = 14
            salto = 31

            for i in range(len(df)):

                fila = fila_base + (i * salto)

                firma = Image(FIRMA_FILE)

                firma.width = 180
                firma.height = 80

                ws.add_image(
                    firma,
                    f"F{fila}"
                )

        # ==========================================
        # GUARDAR EN MEMORIA
        # ==========================================

        excel_buffer = BytesIO()

        wb.save(excel_buffer)

        excel_buffer.seek(0)

        logger.info(
            "fin",
            "Excel generado correctamente",
            {
                "registros": len(df)
            }
        )

        return send_file(

            excel_buffer,

            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

            as_attachment=True,

            download_name=f"Formato_Rechazos_Sura_{fecha_actual.strftime('%d-%m-%Y')}.xlsx"

        )

    except Exception as e:

        logger.error(

            "error",

            "Error general",

            {
                "error": str(e),
                "trace": traceback.format_exc()
            }

        )

        return jsonify({

            "error": str(e)

        }), 500


# ==========================================================
# RUN
# ==========================================================

if __name__ == "__main__":

    app.run(
        host="0.0.0.0",
        port=10000
    )
